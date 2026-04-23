"""Excel file read/write utilities with multi-sheet support."""

from pathlib import Path
from typing import Dict, List, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def save_dataframe_to_excel(df: pd.DataFrame,
                            filepath: Path,
                            sheet_name: str = "Sheet1",
                            index: bool = False) -> None:
    """
    Save a DataFrame to an Excel file with a single sheet.

    Args:
        df: DataFrame to save
        filepath: Output file path
        sheet_name: Name of the sheet
        index: Whether to include index
    """
    filepath.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index)


def save_dataframes_to_excel(dfs: Dict[str, pd.DataFrame],
                            filepath: Path,
                            index: bool = False) -> None:
    """
    Save multiple DataFrames to a single Excel file with multiple sheets.

    Args:
        dfs: Dictionary of {sheet_name: DataFrame}
        filepath: Output file path
        index: Whether to include index
    """
    filepath.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=index)


def append_sheet_to_excel(df: pd.DataFrame,
                         filepath: Path,
                         sheet_name: str,
                         index: bool = False) -> None:
    """
    Append a DataFrame as a new sheet to an existing Excel file.

    Args:
        df: DataFrame to append
        filepath: Existing Excel file path
        sheet_name: Name of the new sheet
        index: Whether to include index
    """
    # Check if file exists
    if not filepath.exists():
        save_dataframe_to_excel(df, filepath, sheet_name, index)
        return

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index)


def format_excel_header(filepath: Path,
                       sheet_name: str,
                       header_bg_color: str = "4472C4",
                       header_font_color: str = "FFFFFF",
                       header_bold: bool = True) -> None:
    """
    Format the header row of an Excel sheet.

    Args:
        filepath: Excel file path
        sheet_name: Sheet to format
        header_bg_color: Background color for header (hex)
        header_font_color: Font color for header (hex)
        header_bold: Whether header font should be bold
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    # Get header row (row 1)
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=header_bg_color, end_color=header_bg_color, fill_type="solid")
        cell.font = Font(color=header_font_color, bold=header_bold)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(filepath)


def auto_adjust_column_width(filepath: Path, sheet_name: Optional[str] = None) -> None:
    """
    Auto-adjust column widths based on content.

    Args:
        filepath: Excel file path
        sheet_name: Specific sheet to adjust (None for all sheets)
    """
    wb = openpyxl.load_workbook(filepath)

    sheets = [wb[sheet_name]] if sheet_name else wb.worksheets

    for ws in sheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)


def create_summary_sheet(filepath: Path,
                        summary_data: Dict[str, List],
                        sheet_name: str = "汇总") -> None:
    """
    Create a summary sheet with key statistics.

    Args:
        filepath: Excel file path
        summary_data: Dictionary with column names as keys and lists as values
        sheet_name: Name for the summary sheet
    """
    df = pd.DataFrame(summary_data)
    append_sheet_to_excel(df, filepath, sheet_name, index=False)
